import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_styled_excel(df):
    """
    Generates a beautifully formatted and styled Excel report using openpyxl.
    Returns the Excel file as a bytes buffer (io.BytesIO) suitable for download in Streamlit.
    """
    # Create an in-memory bytes buffer
    output = io.BytesIO()
    
    # Write DataFrame to openpyxl workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Report"
    
    # Ensure grid lines are visible
    ws.views.sheetView[0].showGridLines = True
    
    # Colors Palette
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Deep Indigo
    high_perf_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid") # Light Green
    med_perf_fill = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid") # Light Yellow
    low_perf_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") # Light Red
    
    # Fonts
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=10, bold=False, color="1F2937")
    category_font_high = Font(name="Segoe UI", size=10, bold=True, color="166534")
    category_font_med = Font(name="Segoe UI", size=10, bold=True, color="854D0E")
    category_font_low = Font(name="Segoe UI", size=10, bold=True, color="991B1B")

    # Border Styling
    thin_border = Side(border_style="thin", color="E5E7EB")
    cell_border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
    
    # Write headers
    headers = list(df.columns)
    ws.append(headers)
    
    # Format Headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = cell_border
    
    # Write and format data rows
    for r_idx, row_data in enumerate(df.values, 2):
        row_values = list(row_data)
        ws.append(row_values)
        
        # Get category index (Performance_Category should be one of the columns)
        try:
            perf_cat = str(df.loc[r_idx - 2, 'Performance_Category'])
        except KeyError:
            perf_cat = "Unknown"
            
        for c_idx in range(1, len(row_values) + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = data_font
            cell.border = cell_border
            
            # Text alignment: left for text, center for numeric/codes
            header_name = headers[c_idx - 1]
            if header_name in ['Name', 'Email']:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif header_name in ['Mid1', 'Mid2', 'Average']:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
            # Row-level highlights or category cell highlights based on Performance Category
            if header_name == 'Performance_Category':
                if perf_cat == 'High Performance':
                    cell.fill = high_perf_fill
                    cell.font = category_font_high
                elif perf_cat == 'Medium Performance':
                    cell.fill = med_perf_fill
                    cell.font = category_font_med
                elif perf_cat == 'Low Performance':
                    cell.fill = low_perf_fill
                    cell.font = category_font_low

    # Auto-fit column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        for cell in col:
            val_str = str(cell.value or '')
            # Handle float display
            if isinstance(cell.value, float):
                val_str = f"{cell.value:.2f}"
            max_len = max(max_len, len(val_str))
            
        # Give padding
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    # Freeze the header row
    ws.freeze_panes = 'A2'
    
    # Save the workbook to output stream
    wb.save(output)
    output.seek(0)
    
    return output
